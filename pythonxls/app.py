from flask import Flask, render_template, request,redirect
from flask_cors import CORS
from openpyxl import load_workbook
from urllib.parse import unquote
import oracledb
import json
import os
import base64
import redis
from datetime import datetime
from celery import Celery


app = Flask(__name__)
redis_client = redis.StrictRedis(host='redis', port=6379, db=1)
simple_app = Celery('worker',
                    broker='amqp://admin:mypass@rabbit:5672',
                    backend='rpc://')
CORS(app)

def create_oracle_connection():
    User = 'AISSAOUI'
    Password = 'AISSAOUI2020*'
    Host = '192.168.0.204'
    Port = '1521'
    Service_name = 'mydbase'

    #Create a connection
    connection = oracledb.connect(user=User, password=Password, host=Host, port=Port, service_name=Service_name)
    return connection

def close_oracle_connection(connection):
    connection.close()

#@app.route('/<string:idtr>')
def hello(idtr):
	connection = create_oracle_connection()
	cur = connection.cursor()
	sqlquery = "SELECT count(*)\
	from DBASE.ENQUETE_EXT WHERE ID_TRAVAIL like :ID_TRAVAIL"
	cur.execute(sqlquery,{'ID_TRAVAIL': idtr})
	count= cur.fetchone()
	return render_template("helloworld.html", keyword="تم تنفيذ المهمة بنجاح", count=count), {"Refresh": f"15; url=http://localhost/s4/appliedprev/{idtr}"}

@app.route('/app/<string:idtr>')
def s4cd(idtr):
	session_id = request.cookies.get('sessionid')
	cache_key = f'username:{session_id}'

	username = redis_client.get(cache_key)
	user = username.decode('utf-8')
	connection = create_oracle_connection()
	cur = connection.cursor()
	sqlquery = "SELECT SE_TRAVAIL.ID_TRAVAIL,SE_TRAVAIL.REF_COUR,SE_TRAVAIL.DATE_COUR,SE_TRAVAIL.OBJ_COUR \
	from DBASE.SE_TRAVAIL WHERE ID_TRAVAIL like :ID_TRAVAIL"
	cur.execute(sqlquery,{'ID_TRAVAIL': idtr})
	travs = cur.fetchone()

	return render_template("index.html", trav_list=travs, idtrav=idtr, username=user)

def DeleteAllData():
	connection = create_oracle_connection()
	cur = connection.cursor()
	sqlquery = "DELETE FROM DBASE.PERSONNE_TEMP_ENQEXT"
	cur.execute(sqlquery)
	connection.commit()
	close_oracle_connection(connection)

@app.route('/app', methods=['POST'])
def submit_form():
	if request.method == 'POST':
		idtr = request.form['idtr']
		username = request.form['username']
		gharadh = request.form['gharadh']
		uploaded_file = request.files['thexcel']

		file_name = uploaded_file.filename
		destination = os.path.join('uploads/', file_name)
		path = 'uploads/'+file_name
		if file_name =='':
			return "Error: Please add an excel file !!"
		uploaded_file.save(destination)
		workbook = load_workbook(filename=path)	
		sheet = workbook.active

		DeleteAllData()

		connection = create_oracle_connection()
		cur = connection.cursor()
		sqlquery = "alter session set nls_date_format = 'dd/mm/rrrr'"
		cur.execute(sqlquery)
		sqlquery1="INSERT INTO DBASE.PERSONNE_TEMP_ENQEXT\
			(CIN, PRENOM, PRENPERE, PRENGPERE, NOM, SEX, NOMPRENMERE, DATNAIS, LIEUNAIS, DATECIN, SITFAM, RESIDENCE, NATIONALITE, PROFESSION, NORDRE)\
			VALUES (:CIN, :PRENOM, :PRENPERE, :PRENGPERE, :NOM, :SEX, :NOMPRENMERE, :DATNAIS, :LIEUNAIS , :DATECIN, :SITFAM, :RESIDENCE, :NATIONALITE, :PROFESSION, :NORDRE)"

		if (sheet["A1"].value == "ع/ر" and sheet["B1"].value == "رقم ب ت و" and sheet["C1"].value == "تاريخ الإصدار" and sheet["D1"].value == "الإسم" 
		and sheet["E1"].value == "إسم الأب" and sheet["F1"].value == "إسم الجد" and sheet["G1"].value == "اللقب" and sheet["H1"].value == "إسم الأم ولقبها"
		and sheet["I1"].value == "تاريخ الولادة" and sheet["J1"].value == "مكان الولادة" and sheet["K1"].value == "الجنس" and sheet["L1"].value == "الجنسية"
		and sheet["M1"].value == "المهنة" and sheet["N1"].value == "الحالة العائلية" and sheet["O1"].value == "العنوان الحالي"):
			for row in sheet.iter_rows(min_row=2, min_col=1, max_col=15, values_only=True):
				nordre = row[0]
				cin = row[1]
				datecin = row[2]
				#datecin = row[2].date().strftime("%d/%m/%Y")
				prenom = row[3]
				prenpere = row[4]
				prengpere = row[5]
				nom = row[6]
				nomprenmere = row[7]
				datnais = row[8]
				#datnais = row[8].date().strftime("%d/%m/%Y")
				lieunais = row[9]
				sex = row[10]
				nationalite = row[11]
				profession = row[12]
				sitfam = row[13]
				residence = row[14]
				try:
					cur.execute(sqlquery1,{'CIN': cin,'PRENOM': prenom,'PRENPERE': prenpere,'PRENGPERE': prengpere,'NOM': nom,'SEX': sex,
            		'NOMPRENMERE': nomprenmere,'DATNAIS': datnais,'LIEUNAIS': lieunais,'DATECIN': datecin,'SITFAM':sitfam,'RESIDENCE':residence,
            		'NATIONALITE':nationalite,'PROFESSION':profession,'NORDRE':nordre})
					connection.commit()
				except oracledb.DatabaseError as e:
					error, = e.args
					code = error.code
					message = error.message
					return render_template("error.html", code=code, message=message)
		else:
			os.remove(path)
			return "error excel not ready for processing"		
		close_oracle_connection(connection)
		idjob= call_method(idtr,username,gharadh)
		return redirect(f"http://localhost/s4/cdjobs")


def call_method(idtr,username,gharadh):
	app.logger.info("Invoking Method ")
	r = simple_app.send_task('tasks.cd_add', kwargs={'idtr': idtr, 'username': username, 'gharadh': gharadh})
	app.logger.info(r.backend)
	connection = create_oracle_connection()
	cur = connection.cursor()

	datejob = datetime.now().strftime("%Y/%m/%d")
	sql = "INSERT INTO CDJOB (THE_USER, DATEJOB, ID_TRAVAIL, CODE, TASK_ID, TASK_STATUS) \
	VALUES(:THE_USER, to_date(:DATEJOB,'YYYY/MM/DD'), :ID_TRAVAIL, :CODE, :TASK_ID, :TASK_STATUS)"
	cur.execute(sql,{'THE_USER':username,'DATEJOB':datejob,'ID_TRAVAIL':idtr,'CODE':gharadh, 'TASK_ID': r.id, 'TASK_STATUS': 'SCHEDULED'})
	connection.commit()
	return r.id

@app.route('/task_status/<task_id>')
def get_status(task_id):
	status = simple_app.AsyncResult(task_id, app=simple_app)
	print("Invoking Method ")
	return str(status.state)

@app.route('/task_result/<task_id>')
def task_result(task_id):
	result = simple_app.AsyncResult(task_id).result
	return "Result of the Task " + str(result)



if __name__ == '__main__':
	app.run(use_reloader=True, host='0.0.0.0', port=5000)

